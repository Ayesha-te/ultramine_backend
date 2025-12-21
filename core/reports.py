from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from django.utils import timezone
from django.db import models
from decimal import Decimal

from .models import Deposit, Withdrawal, Order, DailyEarning
from users.models import User


def generate_users_report_excel():
    """Generate Users Report in Excel format"""
    buffer = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Users Report"
    
    headers = ["ID", "Username", "Email", "Phone", "Registration Date", "Status", "Total Deposits", "Total Earnings"]
    worksheet.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    users = User.objects.all()
    for user in users:
        total_deposits = Deposit.objects.filter(
            user=user, 
            status='approved'
        ).aggregate(total=models.Sum('amount'))['total'] or Decimal('0.00')
        
        total_earnings = DailyEarning.objects.filter(
            user=user
        ).aggregate(total=models.Sum('earning_amount'))['total'] or Decimal('0.00')
        
        worksheet.append([
            user.id,
            user.username,
            user.email,
            user.phone or "N/A",
            user.date_joined.strftime("%Y-%m-%d %H:%M"),
            "Active" if user.is_active else "Inactive",
            f"₨{total_deposits:,.2f}",
            f"₨{total_earnings:,.2f}"
        ])
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def generate_earnings_report_excel():
    """Generate Earnings Report in Excel format"""
    buffer = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Earnings Report"
    
    headers = ["Date", "User", "Email", "Earning Type", "Amount", "Balance"]
    worksheet.append(headers)
    
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    daily_earnings = DailyEarning.objects.all().order_by('-created_at')[:1000]
    for earning in daily_earnings:
        worksheet.append([
            earning.created_at.strftime("%Y-%m-%d %H:%M"),
            earning.user.username,
            earning.user.email,
            "Mining ROI",
            f"₨{earning.earning_amount:,.2f}",
            f"₨{earning.user.wallet.balance:,.2f}"
        ])
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def generate_orders_report_excel():
    """Generate Orders Report in Excel format"""
    buffer = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders Report"
    
    headers = ["Order ID", "Customer", "Email", "Product", "Quantity", "Price", "Total", "Status", "Date"]
    worksheet.append(headers)
    
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    orders = Order.objects.all().order_by('-created_at')
    for order in orders:
        worksheet.append([
            order.id,
            order.customer_name or order.user.username,
            order.email,
            order.product.name,
            order.quantity,
            f"₨{order.product.price:,.2f}",
            f"₨{order.final_price:,.2f}",
            order.status.capitalize(),
            order.created_at.strftime("%Y-%m-%d %H:%M")
        ])
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def generate_users_report_pdf():
    """Generate Users Report in PDF format"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=30,
        alignment=1
    )
    
    elements.append(Paragraph("Users Report", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 0.3 * inch))
    
    users = User.objects.all()
    data = [["Username", "Email", "Phone", "Status", "Joined"]]
    
    for user in users[:50]:
        data.append([
            user.username,
            user.email,
            user.phone or "N/A",
            "Active" if user.is_active else "Inactive",
            user.date_joined.strftime("%Y-%m-%d")
        ])
    
    table = Table(data, colWidths=[1.2*inch, 1.5*inch, 1*inch, 0.8*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_earnings_report_pdf():
    """Generate Earnings Report in PDF format"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#70AD47'),
        spaceAfter=30,
        alignment=1
    )
    
    elements.append(Paragraph("Earnings Report", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 0.3 * inch))
    
    daily_earnings = DailyEarning.objects.all().order_by('-created_at')[:50]
    data = [["Date", "User", "Amount", "Balance"]]
    
    for earning in daily_earnings:
        data.append([
            earning.created_at.strftime("%Y-%m-%d"),
            earning.user.username,
            f"₨{earning.earning_amount:,.2f}",
            f"₨{earning.user.wallet.balance:,.2f}"
        ])
    
    table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_orders_report_pdf():
    """Generate Orders Report in PDF format"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#FFC000'),
        spaceAfter=30,
        alignment=1
    )
    
    elements.append(Paragraph("Orders Report", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 0.3 * inch))
    
    orders = Order.objects.all().order_by('-created_at')[:50]
    data = [["Order ID", "Customer", "Product", "Qty", "Total", "Status", "Date"]]
    
    for order in orders:
        data.append([
            str(order.id),
            order.customer_name or order.user.username,
            order.product.name[:20],
            str(order.quantity),
            f"₨{order.final_price:,.2f}",
            order.status.capitalize(),
            order.created_at.strftime("%Y-%m-%d")
        ])
    
    table = Table(data, colWidths=[0.8*inch, 1.2*inch, 1.2*inch, 0.6*inch, 1.2*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFC000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer
